from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
wb = load_workbook(filename="1.xlsx")
b=wb.active
c=b.cell(1,1).value
d=b.cell(2,1).value#zelle auslesen
b.cell(2,1).value=56#Zelle änder
i=0
while i<10:
    i=i+1
    b.cell(i,1).value=i

a1 = b['A1']
gr = Font(color=colors.GREEN)#Farbe setzen
rt = Font(color=colors.RED)#Farbe setzen
randomcolor= Font(color='001280')
a1.font=randomcolor
print(a1.value)
wb.save("1.xlsx")